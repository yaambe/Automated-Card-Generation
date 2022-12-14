import openpyxl as op
import win32com.client
import os
# initial data in the excel sheets

fullname_to_find = 'Ahmed Yameen'
designation_to_find = 'Senior Superintendent'
mobile_to_find = '9916696'
phone_to_find = '7497851'
email_to_find = 'ahmed.yameen'

# The data we are going to capture
fullname = input("Enter Staff Name (*best length is below 15 characters): ").title()
designation = input("Enter Designation: ").title()
mobile = input("Mobile: (960) ")
phone = input("Phone: (960) ")
email = input("Customs E-mail (part before @customs.gov.mv) (*best length is below 15 characters): ")

# editing the template
wb = op.load_workbook(r"template\excel template.xlsx")
ws = wb['pages']
i = 0
for r in range(1 , ws.max_row + 1):
    for c in range(1 , ws.max_column + 1):
        s = ws.cell(r,c).value
        if s != None and fullname_to_find in s: 
            ws.cell(r,c).value = s.replace(fullname_to_find,fullname)
        if s != None and designation_to_find in s: 
            ws.cell(r,c).value = s.replace(designation_to_find,designation)
        if s != None and mobile_to_find in s: 
            ws.cell(r,c).value = s.replace(mobile_to_find,mobile)
        if s != None and phone_to_find in s: 
            ws.cell(r,c).value = s.replace(phone_to_find,phone)
        if s != None and email_to_find in s: 
            ws.cell(r,c).value = s.replace(email_to_find,email)
            i += 1

# All changes have been made

wb.save(f'editables\\{fullname}.xlsx')
wb.close()

# save a copy
o = win32com.client.Dispatch("Excel.Application")
o.Visible = False
wb_path = f'F:\\Git\\Automated-Card-Generation\\editables\\{fullname}.xlsx'

new = o.Workbooks.Open(wb_path)

# open the copy to create a pdf

ws_index_list = [1]


path_to_pdf =  f'F:\\Git\\Automated-Card-Generation\\PDFs\\{fullname}.pdf'
new.WorkSheets(ws_index_list).Select()
new.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)





